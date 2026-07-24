#!/usr/bin/env python3
"""
Censo de vacantes de liquidación en transporte — Indeed MX + Computrabajo.

Versión reforzada:
- Revisión una por una: ninguna vacante se clasifica sin descripción COMPLETA.
  Las que no se logren verificar van a la hoja "Sin verificar" para reintento.
- Módulo propio de Computrabajo con Playwright (paginación completa + visita
  individual a cada vacante + screenshot de evidencia de las que son SEÑAL).
- Unificación de fuentes en un solo Excel con dedup fuzzy (thefuzz, umbral 85);
  si la misma vacante está en ambas se conserva la versión de Computrabajo y
  la fuente se marca como "ambas".
- Checkpoint incremental cada 25 vacantes; al reanudar salta lo ya descargado.

Uso:
    python scraper_liquidacion.py                          # corrida completa (90 días)
    python scraper_liquidacion.py --hours-old 24 --append  # tarea diaria incremental
    python scraper_liquidacion.py --headful                # para resolver captcha a mano
    python scraper_liquidacion.py --test                   # valida el pipeline sin red
"""

import argparse
import logging
import random
import re
import sys
import time
import unicodedata
from datetime import datetime
from pathlib import Path
from urllib.parse import quote, urljoin

import pandas as pd

# ----------------------------------------------------------------------------
# Configuración de búsquedas
# ----------------------------------------------------------------------------

KEYWORDS = [
    "liquidador",
    "cajero liquidador",
    "liquidación de rutas",
    "liquidador de viajes",
    "ejecutivo de liquidaciones",
    "analista de liquidaciones",
    "comprobación de gastos operadores",
    "gastos de viaje transporte",
    "analista de combustible",
    "control de diésel",
    "carta porte",
    "facturista transporte",
    "cuadre de rutas",
    "póliza de gastos transporte",
]

# (parámetro location para jobspy, estado para reportes)
LOCATIONS = [
    ("Ciudad de México", "Ciudad de México"),
    ("Estado de México", "Estado de México"),
    ("Monterrey NL", "Nuevo León"),
    ("Guadalajara Jal", "Jalisco"),
    ("León Gto", "Guanajuato"),
    ("Querétaro Qro", "Querétaro"),
    ("Puebla Pue", "Puebla"),
    ("Veracruz Ver", "Veracruz"),
    ("Mérida Yuc", "Yucatán"),
    ("Tijuana BC", "Baja California"),
    ("Culiacán Sin", "Sinaloa"),
    ("Torreón Coah", "Coahuila"),
]

MIN_DESC_COMPLETA = 300  # descripción menor a esto se considera truncada

CT_BASE = "https://mx.computrabajo.com"
CT_USER_AGENT = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                 "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36")

# ----------------------------------------------------------------------------
# Reglas de clasificación (se evalúan sobre título + descripción en minúsculas)
# ----------------------------------------------------------------------------

SIGNAL_WORDS = [
    "ruta", "rutas", "chofer", "choferes", "repartidor", "reparto",
    "operador", "operadores", "viaje", "viajes", "unidades", "flota",
    "cedis", "porteo", "folios con choferes", "diésel", "diesel",
    "casetas", "carta porte",
]

# Cada regla de ruido: (nombre del motivo, palabras que la disparan)
NOISE_RULES = [
    ("liquidación de nómina / finiquitos",
     ["finiquito", "nómina", "imss patronal", "indemnización laboral"]),
    ("cajero de mostrador / ventanilla",
     ["piso de venta", "mostrador", "punto de venta", "tienda departamental",
      "autoservicio"]),
    ("cobranza domiciliaria de servicios",
     ["servicio suspendido", "reconexión", "recuperación de equipos",
      "telecomunicaciones"]),
    ("traslado de valores como giro",
     ["traslado de valores", "custodia de valores"]),
    ("seguros (siniestros / ajustador)",
     ["liquidador de siniestros", "ajustador", "siniestros"]),
]

# "billetes falsos" solo es ruido si NO hay mención de rutas
NOISE_BILLETES = ("cajero de mostrador / ventanilla", ["billetes falsos"])

# Empresas que parecen agencia de reclutamiento: se marcan, no se descartan
AGENCY_WORDS = [
    "rh", "recursos humanos", "reclutamiento", "staffing",
    "consultores", "capital humano",
]

log = logging.getLogger("censo")


def normalize(text) -> str:
    if text is None or (isinstance(text, float) and pd.isna(text)):
        return ""
    return str(text).lower().strip()


def strip_accents(text: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", text)
        if unicodedata.category(c) != "Mn"
    )


def slugify(text: str) -> str:
    plain = strip_accents(normalize(text))
    return re.sub(r"[^a-z0-9]+", "-", plain).strip("-")


def find_words(text: str, words) -> list:
    """Palabras presentes en el texto, con frontera de palabra y sin acentos."""
    plain = strip_accents(text)
    hits = []
    for w in words:
        pattern = r"(?<![\wáéíóúñ])" + re.escape(strip_accents(w)) + r"(?![\wáéíóúñ])"
        if re.search(pattern, plain):
            hits.append(w)
    return hits


def classify(title: str, description: str, company: str) -> dict:
    """Devuelve clasificación SEÑAL / RUIDO / REVISAR con detalle de reglas."""
    text = normalize(title) + " " + normalize(description)
    signal_hits = find_words(text, SIGNAL_WORDS)

    noise_hits = []  # [(motivo, palabra), ...]
    for motivo, words in NOISE_RULES:
        for w in find_words(text, words):
            noise_hits.append((motivo, w))
    # billetes falsos: solo cuenta como ruido sin mención de rutas
    if not find_words(text, ["ruta", "rutas"]):
        for w in find_words(text, NOISE_BILLETES[1]):
            noise_hits.append((NOISE_BILLETES[0], w))

    es_agencia = bool(find_words(normalize(company), AGENCY_WORDS))

    if signal_hits and noise_hits:
        clasificacion = "REVISAR"
    elif signal_hits:
        clasificacion = "SEÑAL"
    else:
        clasificacion = "RUIDO"

    motivos = "; ".join(sorted({f"{m} ({w})" for m, w in noise_hits}))
    if clasificacion == "RUIDO" and not noise_hits:
        motivos = "sin palabras de señal"

    return {
        "clasificacion": clasificacion,
        "palabras_senal": ", ".join(signal_hits),
        "motivo_descarte": motivos,
        "es_agencia": "AGENCIA" if es_agencia else "",
    }


# ----------------------------------------------------------------------------
# Checkpoint incremental (resume)
# ----------------------------------------------------------------------------

class Checkpoint:
    """CSV incremental: cada N vacantes nuevas se persisten a disco.

    Al reanudar una corrida interrumpida, `seen_urls` permite saltar
    las vacantes ya descargadas.
    """

    def __init__(self, path: Path, every: int = 25):
        self.path = Path(path)
        self.every = every
        self.buffer = []
        self.seen_urls = set()
        if self.path.exists():
            try:
                prev = pd.read_csv(self.path)
                self.seen_urls = set(prev.get("job_url", pd.Series(dtype=str)).dropna())
                log.info("Checkpoint existente: %d vacantes ya descargadas (se saltarán)",
                         len(self.seen_urls))
            except Exception as exc:  # noqa: BLE001
                log.warning("No se pudo leer checkpoint %s: %s", self.path, exc)

    def add(self, rows):
        if isinstance(rows, pd.DataFrame):
            rows = rows.to_dict("records")
        elif isinstance(rows, dict):
            rows = [rows]
        for r in rows:
            url = str(r.get("job_url") or "")
            if url:
                self.seen_urls.add(url)
            self.buffer.append(r)
        if len(self.buffer) >= self.every:
            self.flush()

    def flush(self):
        if not self.buffer:
            return
        df = pd.DataFrame(self.buffer)
        header = not self.path.exists()
        df.to_csv(self.path, mode="a", header=header, index=False)
        log.info("Checkpoint: %d vacantes persistidas en %s", len(self.buffer), self.path)
        self.buffer = []

    def load_all(self) -> pd.DataFrame:
        self.flush()
        if self.path.exists():
            return pd.read_csv(self.path)
        return pd.DataFrame()


# ----------------------------------------------------------------------------
# Fuente 1: Indeed (fallback LinkedIn) vía jobspy
# ----------------------------------------------------------------------------

def scrape_one(keyword: str, location: str, results_wanted: int, hours_old: int) -> pd.DataFrame:
    """Una búsqueda: Indeed primero, LinkedIn como fuente secundaria."""
    from jobspy import scrape_jobs

    last_error = None
    for site in ("indeed", "linkedin"):
        try:
            kwargs = dict(
                site_name=site,
                search_term=keyword,
                location=location,
                results_wanted=results_wanted,
                hours_old=hours_old,
                description_format="markdown",
                linkedin_fetch_description=(site == "linkedin"),
            )
            if site == "indeed":
                kwargs["country_indeed"] = "Mexico"
            df = scrape_jobs(**kwargs)
            if df is not None and len(df) > 0:
                df["fuente"] = site
                return df
            log.info("  %s sin resultados para '%s' en '%s'", site, keyword, location)
        except Exception as exc:  # noqa: BLE001 — cualquier fallo pasa al siguiente sitio
            last_error = exc
            log.warning("  %s falló para '%s' en '%s': %s", site, keyword, location, exc)
    if last_error is not None:
        raise last_error
    return pd.DataFrame()


def run_jobspy_searches(results_wanted, hours_old, min_sleep, max_sleep, checkpoint):
    frames, ok, failed = [], 0, 0
    total = len(KEYWORDS) * len(LOCATIONS)
    n = 0
    for keyword in KEYWORDS:
        for location, estado in LOCATIONS:
            n += 1
            log.info("[jobspy %d/%d] '%s' en '%s'", n, total, keyword, location)
            try:
                df = scrape_one(keyword, location, results_wanted, hours_old)
                if len(df) > 0:
                    df["busqueda_keyword"] = keyword
                    df["busqueda_plaza"] = location
                    df["busqueda_estado"] = estado
                    df = df[~df["job_url"].astype(str).isin(checkpoint.seen_urls)]
                    if len(df) > 0:
                        checkpoint.add(df)
                        frames.append(df)
                ok += 1
            except Exception as exc:  # noqa: BLE001 — registrar y continuar
                failed += 1
                log.error("BÚSQUEDA FALLIDA (jobspy) '%s' en '%s': %s", keyword, location, exc)
            if n < total:
                time.sleep(random.uniform(min_sleep, max_sleep))
    raw = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
    return raw, ok, failed


# ----------------------------------------------------------------------------
# Fuente 2: Computrabajo vía Playwright (una por una, con evidencia)
# ----------------------------------------------------------------------------

class CaptchaError(RuntimeError):
    pass


def launch_chromium(pw, headless: bool):
    """Lanza Chromium; si la revisión descargada no coincide, usa el ejecutable
    del sistema (CHROMIUM_PATH o /opt/pw-browsers/chromium)."""
    import os
    try:
        return pw.chromium.launch(headless=headless)
    except Exception as exc:  # noqa: BLE001
        for candidate in (os.environ.get("CHROMIUM_PATH"), "/opt/pw-browsers/chromium",
                          "/usr/bin/chromium", "/usr/bin/chromium-browser"):
            if candidate and Path(candidate).exists():
                log.info("Chromium de Playwright no disponible (%s); usando %s", exc, candidate)
                return pw.chromium.launch(headless=headless, executable_path=candidate)
        raise


class ComputrabajoScraper:
    """Busca las keywords en Computrabajo, pagina hasta el final y visita
    CADA vacante individualmente para extraer la descripción completa."""

    CARD_SELECTORS = ["article.box_offer", "article[data-id]", "article"]
    TITLE_SELECTORS = ["h2 a", "h1 a", "a.js-o-link"]
    DETAIL_DESC_SELECTORS = [
        "div[class*='detail'] p", ".box_detail", "[class*='description']",
        "main p",
    ]

    def __init__(self, headful=False, storage_state="ct_storage_state.json",
                 capturas_dir="capturas", min_sleep=4.0, max_sleep=10.0,
                 seen_urls=None, max_pages=60):
        self.headful = headful
        self.storage_state = Path(storage_state)
        self.capturas_dir = Path(capturas_dir)
        self.min_sleep = min_sleep
        self.max_sleep = max_sleep
        self.seen_urls = seen_urls if seen_urls is not None else set()
        self.max_pages = max_pages
        self._pw = None
        self._browser = None
        self._context = None
        self.page = None

    # -- ciclo de vida -------------------------------------------------------

    def start(self):
        from playwright.sync_api import sync_playwright
        self._pw = sync_playwright().start()
        self._browser = launch_chromium(self._pw, headless=not self.headful)
        ctx_kwargs = dict(user_agent=CT_USER_AGENT, locale="es-MX",
                          viewport={"width": 1366, "height": 900})
        if self.storage_state.exists():
            ctx_kwargs["storage_state"] = str(self.storage_state)
            log.info("Computrabajo: reusando cookies de %s", self.storage_state)
        self._context = self._browser.new_context(**ctx_kwargs)
        self.page = self._context.new_page()
        self.capturas_dir.mkdir(parents=True, exist_ok=True)

    def stop(self):
        try:
            if self._context:
                self._context.storage_state(path=str(self.storage_state))
                self._context.close()
            if self._browser:
                self._browser.close()
            if self._pw:
                self._pw.stop()
        except Exception:  # noqa: BLE001
            pass

    def _sleep(self):
        time.sleep(random.uniform(self.min_sleep, self.max_sleep))

    # -- captcha / cloudflare -------------------------------------------------

    def _check_captcha(self):
        title = normalize(self.page.title())
        challenge = self.page.locator(
            "#challenge-form, #cf-chl-widget, iframe[src*='challenge'], "
            "iframe[src*='captcha'], #turnstile-wrapper").count()
        if challenge == 0 and "un momento" not in title and "just a moment" not in title:
            return
        if not self.headful:
            raise CaptchaError(
                "Computrabajo pidió captcha/Cloudflare. Reintenta con --headful, "
                "resuélvelo una vez y las cookies se guardarán para el resto.")
        log.warning("CAPTCHA detectado: resuélvelo en la ventana del navegador "
                    "(esperando hasta 5 minutos)...")
        deadline = time.time() + 300
        while time.time() < deadline:
            time.sleep(3)
            title = normalize(self.page.title())
            if ("un momento" not in title and "just a moment" not in title
                    and self.page.locator("#challenge-form").count() == 0):
                self._context.storage_state(path=str(self.storage_state))
                log.info("Captcha resuelto; cookies guardadas en %s", self.storage_state)
                return
        raise CaptchaError("Captcha no resuelto en 5 minutos; abortando Computrabajo")

    def _goto(self, url: str):
        self.page.goto(url, wait_until="domcontentloaded", timeout=60000)
        self._check_captcha()

    # -- listado --------------------------------------------------------------

    def search_keyword(self, keyword: str) -> list:
        """Regresa [{titulo, empresa, ciudad, job_url}] de todas las páginas."""
        slug = slugify(keyword)
        base_urls = [
            f"{CT_BASE}/trabajo-de-{slug}",
            f"{CT_BASE}/trabajo-de-{slug.split('-')[0]}?q={quote(keyword)}",
        ]
        cards, last_error = [], None
        for base in base_urls:
            try:
                cards = self._paginate(base, keyword)
            except CaptchaError:
                raise
            except Exception as exc:  # noqa: BLE001
                last_error = exc
                continue
            if cards:
                return cards
        if not cards and last_error is not None:
            raise last_error
        return cards

    def _paginate(self, base_url: str, keyword: str) -> list:
        results = []
        for p in range(1, self.max_pages + 1):
            sep = "&" if "?" in base_url else "?"
            url = base_url if p == 1 else f"{base_url}{sep}p={p}"
            try:
                self._goto(url)
            except CaptchaError:
                raise
            except Exception as exc:  # noqa: BLE001
                if p == 1:  # ni la primera página cargó: búsqueda fallida
                    raise
                log.warning("  Computrabajo página %d de '%s' falló: %s", p, keyword, exc)
                break
            page_cards = self._extract_cards()
            if not page_cards:
                break
            new = [c for c in page_cards if c["job_url"] not in {r["job_url"] for r in results}]
            if not new:  # la paginación regresó lo mismo: fin
                break
            results.extend(new)
            log.info("  Computrabajo '%s' página %d: %d vacantes (acumulado %d)",
                     keyword, p, len(new), len(results))
            self._sleep()
        return results

    def _extract_cards(self) -> list:
        cards = []
        for sel in self.CARD_SELECTORS:
            articles = self.page.locator(sel)
            if articles.count() > 0:
                break
        else:
            return cards
        for i in range(articles.count()):
            art = articles.nth(i)
            try:
                link = None
                for tsel in self.TITLE_SELECTORS:
                    loc = art.locator(tsel)
                    if loc.count() > 0:
                        link = loc.first
                        break
                if link is None:
                    continue
                href = link.get_attribute("href") or ""
                if not href:
                    continue
                titulo = (link.text_content() or "").strip()
                texto = [t.strip() for t in (art.text_content() or "").split("\n") if t.strip()]
                empresa, ciudad = "", ""
                comp_loc = art.locator("p a, .fs16 a")
                if comp_loc.count() > 0:
                    empresa = (comp_loc.first.text_content() or "").strip()
                loc_loc = art.locator("p span, [class*='location']")
                if loc_loc.count() > 0:
                    ciudad = (loc_loc.first.text_content() or "").strip()
                if not empresa and len(texto) > 1:
                    empresa = texto[1]
                cards.append({
                    "titulo": titulo, "empresa": empresa, "ciudad": ciudad,
                    "job_url": urljoin(CT_BASE, href),
                })
            except Exception:  # noqa: BLE001 — tarjeta ilegible, seguir
                continue
        return cards

    # -- detalle (una por una) -------------------------------------------------

    def fetch_detail(self, card: dict) -> dict:
        """Entra a la vacante y extrae descripción completa + metadatos."""
        self._goto(card["job_url"])
        page_text = self.page.locator("body").inner_text(timeout=15000)

        descripcion = ""
        for sel in self.DETAIL_DESC_SELECTORS:
            loc = self.page.locator(sel)
            if loc.count() > 0:
                descripcion = "\n".join(
                    (loc.nth(i).text_content() or "").strip()
                    for i in range(min(loc.count(), 40)))
                if len(descripcion) >= 100:
                    break
        if len(descripcion) < 100:
            descripcion = page_text  # último recurso: texto completo de la página

        sueldo_min = sueldo_max = None
        m = re.search(r"\$\s?([\d.,]+)(?:\s*(?:a|-|hasta)\s*\$?\s?([\d.,]+))?", page_text)
        if m:
            def num(s):
                try:
                    return float(s.replace(",", "").replace(".00", ""))
                except ValueError:
                    return None
            sueldo_min = num(m.group(1))
            sueldo_max = num(m.group(2)) if m.group(2) else None

        contrato = ""
        m = re.search(r"contrato[^\n]{0,60}", page_text, re.IGNORECASE)
        if m:
            contrato = m.group(0).strip()

        fecha = ""
        m = re.search(r"(publicad[ao][^\n]{0,40}|hace \d+[^\n]{0,20}|\d{1,2} de \w+ de \d{4})",
                      page_text, re.IGNORECASE)
        if m:
            fecha = m.group(0).strip()

        empresa = card["empresa"]
        emp_loc = self.page.locator("[class*='company'] a, .container a[href*='empresa'], h2 a")
        if not empresa and emp_loc.count() > 0:
            empresa = (emp_loc.first.text_content() or "").strip()

        return {
            "title": card["titulo"],
            "company": empresa,
            "location": card["ciudad"],
            "description": descripcion,
            "min_amount": sueldo_min,
            "max_amount": sueldo_max,
            "tipo_contrato": contrato,
            "date_posted": fecha,
            "job_url": card["job_url"],
            "fuente": "computrabajo",
        }

    def screenshot(self, empresa: str, puesto: str) -> str:
        fecha = datetime.now().strftime("%Y-%m-%d")
        name = re.sub(r"[^\w\-]+", "_", strip_accents(f"{empresa}_{puesto}"))[:100]
        path = self.capturas_dir / f"{name}_{fecha}.png"
        try:
            self.page.screenshot(path=str(path), full_page=True)
            return str(path)
        except Exception as exc:  # noqa: BLE001
            log.warning("No se pudo capturar screenshot de %s: %s", empresa, exc)
            return ""

    # -- corrida completa --------------------------------------------------------

    def run(self, keywords, checkpoint: Checkpoint):
        """Regresa (dataframe, búsquedas_ok, búsquedas_fallidas)."""
        rows, ok, failed = [], 0, 0
        for k, keyword in enumerate(keywords, 1):
            log.info("[computrabajo %d/%d] '%s'", k, len(keywords), keyword)
            try:
                cards = self.search_keyword(keyword)
                log.info("  %d vacantes listadas para '%s'", len(cards), keyword)
                for card in cards:
                    if card["job_url"] in checkpoint.seen_urls:
                        continue
                    try:
                        detail = self.fetch_detail(card)
                        detail["busqueda_keyword"] = keyword
                        detail["busqueda_plaza"] = "Computrabajo (nacional)"
                        detail["busqueda_estado"] = ""
                        c = classify(detail["title"], detail["description"],
                                     detail["company"])
                        if c["clasificacion"] == "SEÑAL":
                            detail["screenshot"] = self.screenshot(
                                detail["company"], detail["title"])
                        rows.append(detail)
                        checkpoint.add(detail)
                    except CaptchaError:
                        raise
                    except Exception as exc:  # noqa: BLE001
                        log.error("VACANTE FALLIDA %s: %s", card["job_url"], exc)
                    self._sleep()
                ok += 1
            except CaptchaError as exc:
                log.error("Computrabajo detenido por captcha: %s", exc)
                failed += len(keywords) - k + 1
                break
            except Exception as exc:  # noqa: BLE001
                failed += 1
                log.error("BÚSQUEDA FALLIDA (computrabajo) '%s': %s", keyword, exc)
        return pd.DataFrame(rows), ok, failed


# ----------------------------------------------------------------------------
# Descripciones completas para filas de jobspy (revisión una por una)
# ----------------------------------------------------------------------------

def is_desc_complete(desc) -> bool:
    return len(normalize(desc)) >= MIN_DESC_COMPLETA


def fetch_missing_descriptions(df: pd.DataFrame, headful: bool) -> pd.DataFrame:
    """Para filas jobspy con descripción vacía/truncada, visita el job_url
    individual con Playwright y rellena la descripción completa."""
    df = df.copy()
    if "description" not in df.columns:
        df["description"] = None
    df["desc_verificada"] = df["description"].map(is_desc_complete)
    pending = df.index[~df["desc_verificada"]].tolist()
    if not pending:
        return df

    log.info("Revisión una por una: %d vacantes de jobspy sin descripción completa; "
             "visitando su job_url individual...", len(pending))
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as pw:
            browser = launch_chromium(pw, headless=not headful)
            ctx = browser.new_context(user_agent=CT_USER_AGENT, locale="es-MX")
            page = ctx.new_page()
            for idx in pending:
                url = str(df.at[idx, "job_url"] or "")
                if not url:
                    continue
                try:
                    page.goto(url, wait_until="domcontentloaded", timeout=45000)
                    desc_loc = page.locator(
                        "#jobDescriptionText, [class*='jobsearch-JobComponent-description'], "
                        ".description__text, [class*='description']")
                    if desc_loc.count() > 0:
                        text = desc_loc.first.inner_text(timeout=10000).strip()
                    else:
                        text = page.locator("body").inner_text(timeout=10000).strip()
                    if len(text) > len(normalize(df.at[idx, "description"])):
                        df.at[idx, "description"] = text
                    # se visitó la página real: la descripción ya es la completa,
                    # aunque el puesto tenga una descripción genuinamente corta
                    df.at[idx, "desc_verificada"] = len(text) >= 100
                except Exception as exc:  # noqa: BLE001
                    log.warning("  No se pudo completar descripción de %s: %s", url, exc)
                time.sleep(random.uniform(2, 5))
            browser.close()
    except Exception as exc:  # noqa: BLE001
        log.error("Playwright no disponible para completar descripciones: %s", exc)
    return df


# ----------------------------------------------------------------------------
# Datos simulados para validar el pipeline sin red (--test)
# ----------------------------------------------------------------------------

def mock_jobs() -> pd.DataFrame:
    LONG = (" Responsabilidades: recepción y arqueo diario, revisión de comprobantes, "
            "conciliación contra sistema, reporte a gerencia administrativa, manejo de "
            "efectivo y valores, atención a auditorías internas y trato directo con el "
            "personal operativo de la empresa en sucursal y oficina central.")
    rows = [
        # señal clara (indeed)
        dict(title="Liquidador de rutas", company="Transportes del Norte",
             location="Monterrey", description="Liquidación de rutas foráneas, cuadre de casetas y diésel con operadores. Manejo de carta porte." + LONG,
             min_amount=12000, max_amount=15000, fuente="indeed",
             busqueda_keyword="liquidador", busqueda_plaza="Monterrey NL", busqueda_estado="Nuevo León"),
        # duplicado fuzzy en computrabajo (misma empresa/ciudad, título 'similar'):
        # debe conservarse esta versión y marcarse fuente = "ambas"
        dict(title="Liquidador de Rutas Foráneas", company="Transportes del Norte",
             location="Monterrey", description="Liquidación de rutas foráneas: cuadre de casetas, diésel y viáticos de operadores. Carta porte y folios con choferes." + LONG,
             min_amount=12500, max_amount=15500, fuente="computrabajo", tipo_contrato="Contrato por tiempo indeterminado",
             busqueda_keyword="liquidación de rutas", busqueda_plaza="Computrabajo (nacional)", busqueda_estado=""),
        # señal (computrabajo exclusiva)
        dict(title="Cajero liquidador", company="Logística Azteca",
             location="Ciudad de México", description="Recepción de efectivo de choferes al cierre de reparto, arqueo de folios con choferes." + LONG,
             min_amount=10000, max_amount=None, fuente="computrabajo", tipo_contrato="Contrato indeterminado",
             busqueda_keyword="cajero liquidador", busqueda_plaza="Computrabajo (nacional)", busqueda_estado=""),
        # misma empresa, segunda vacante distinta (vacantes_simultaneas = 2)
        dict(title="Analista de liquidaciones", company="Transportes del Norte",
             location="Monterrey", description="Comprobación de gastos de viaje de operadores, control de flota y unidades." + LONG,
             min_amount=None, max_amount=None, fuente="indeed",
             busqueda_keyword="analista de liquidaciones", busqueda_plaza="Monterrey NL", busqueda_estado="Nuevo León"),
        # ruido: nómina
        dict(title="Analista de liquidaciones", company="Corporativo Fiscal SA",
             location="Guadalajara", description="Cálculo de finiquito e indemnización laboral, nómina semanal, IMSS patronal." + LONG,
             min_amount=None, max_amount=None, fuente="indeed",
             busqueda_keyword="analista de liquidaciones", busqueda_plaza="Guadalajara Jal", busqueda_estado="Jalisco"),
        # ruido: mostrador
        dict(title="Cajero liquidador", company="Tiendas Delta",
             location="Puebla", description="Atención en piso de venta y punto de venta, detección de billetes falsos en mostrador." + LONG,
             min_amount=8000, max_amount=9000, fuente="computrabajo",
             busqueda_keyword="cajero liquidador", busqueda_plaza="Computrabajo (nacional)", busqueda_estado=""),
        # caso borde: señal + ruido → REVISAR
        dict(title="Liquidador", company="Distribuidora del Golfo",
             location="Veracruz", description="Liquidación de rutas de reparto y apoyo en nómina y finiquito del personal." + LONG,
             min_amount=11000, max_amount=13000, fuente="indeed",
             busqueda_keyword="liquidador", busqueda_plaza="Veracruz Ver", busqueda_estado="Veracruz"),
        # agencia con señal → marcar AGENCIA, no descartar
        dict(title="Liquidador de viajes", company="Capital Humano Selecto",
             location="Querétaro", description="Importante empresa de transporte solicita liquidador de viajes, cuadre de diésel y casetas de operadores." + LONG,
             min_amount=13000, max_amount=16000, fuente="computrabajo",
             busqueda_keyword="liquidador de viajes", busqueda_plaza="Computrabajo (nacional)", busqueda_estado=""),
        # ruido: seguros
        dict(title="Liquidador de siniestros", company="Seguros Peninsular",
             location="Mérida", description="Ajustador y liquidador de siniestros automotrices, valuación de daños." + LONG,
             min_amount=None, max_amount=None, fuente="indeed",
             busqueda_keyword="liquidador", busqueda_plaza="Mérida Yuc", busqueda_estado="Yucatán"),
        # descripción truncada (<300) y sin forma de completarla → "Sin verificar"
        dict(title="Ejecutivo de liquidaciones", company="Fletes Rápidos SA",
             location="Tijuana", description="Liquidación de operadores.",
             min_amount=None, max_amount=None, fuente="indeed",
             busqueda_keyword="ejecutivo de liquidaciones", busqueda_plaza="Tijuana BC", busqueda_estado="Baja California"),
    ]
    df = pd.DataFrame(rows)
    df["job_url"] = "https://ejemplo.mx/vacante/" + df.index.astype(str)
    df["date_posted"] = "2026-07-20"
    df["desc_verificada"] = df["description"].map(is_desc_complete)
    return df


# ----------------------------------------------------------------------------
# Unificación y procesamiento
# ----------------------------------------------------------------------------

SOURCE_PRIORITY = {"computrabajo": 0, "ambas": 0, "indeed": 1, "linkedin": 2, "mock": 3}


def extract_city(row) -> str:
    loc = normalize(row.get("location"))
    if loc:
        return str(row.get("location")).split(",")[0].strip()
    return str(row.get("busqueda_plaza", ""))


def city_to_estado(ciudad: str, fallback: str) -> str:
    if fallback:
        return fallback
    plain = strip_accents(normalize(ciudad))
    for loc, estado in LOCATIONS:
        if strip_accents(normalize(loc.split()[0])) in plain or \
           strip_accents(normalize(estado)) in plain:
            return estado
    return ""


def process(raw: pd.DataFrame) -> pd.DataFrame:
    """Clasifica (solo verificadas), deduplica exacto + fuzzy entre fuentes."""
    df = raw.copy()
    for col in ["title", "company", "description", "min_amount", "max_amount",
                "date_posted", "job_url", "location", "fuente", "tipo_contrato",
                "screenshot", "desc_verificada"]:
        if col not in df.columns:
            df[col] = None
    df["desc_verificada"] = df["desc_verificada"].fillna(False).astype(bool)

    records = []
    for _, row in df.iterrows():
        verificada = bool(row["desc_verificada"])
        c = classify(row["title"], row["description"], row["company"])
        desc = normalize(row["description"])
        records.append({
            "empresa": str(row["company"] or "").strip() or "(sin empresa)",
            "titulo": str(row["title"] or "").strip(),
            "ciudad": extract_city(row),
            "estado": city_to_estado(extract_city(row), row.get("busqueda_estado", "") or ""),
            "sueldo_min": row["min_amount"],
            "sueldo_max": row["max_amount"],
            "fecha_publicacion": str(row["date_posted"] or ""),
            "tipo_contrato": str(row["tipo_contrato"] or ""),
            "extracto": desc[:500],
            "palabras_senal": c["palabras_senal"],
            "motivo_descarte": c["motivo_descarte"],
            "liga": str(row["job_url"] or ""),
            "keyword_busqueda": row.get("busqueda_keyword", ""),
            "plaza_busqueda": row.get("busqueda_plaza", ""),
            "fuente": str(row["fuente"] or ""),
            "desc_completa": "sí" if verificada else "no",
            "clasificacion": c["clasificacion"] if verificada else "SIN VERIFICAR",
            "agencia": c["es_agencia"],
            "screenshot": str(row["screenshot"] or ""),
        })
    out = pd.DataFrame(records)
    if out.empty:
        return out

    # 1) deduplicación exacta por empresa + título + ciudad
    out["_key"] = (
        out["empresa"].map(normalize).map(strip_accents) + "|" +
        out["titulo"].map(normalize).map(strip_accents) + "|" +
        out["ciudad"].map(normalize).map(strip_accents)
    )
    out = out.sort_values(by="fuente", key=lambda s: s.map(lambda f: SOURCE_PRIORITY.get(f, 9)))
    dup_exact = out.duplicated(subset="_key", keep="first")
    exact_sources = out.groupby("_key")["fuente"].agg(set)
    out = out[~dup_exact].reset_index(drop=True)

    # 2) deduplicación fuzzy ENTRE fuentes: misma empresa+ciudad, título similar (≥85)
    out = fuzzy_dedup(out, exact_sources)

    # vacantes simultáneas: vacantes distintas por empresa (señal + revisar)
    activas = out[out["clasificacion"].isin(["SEÑAL", "REVISAR"])]
    conteo = activas.groupby(activas["empresa"].map(normalize)).size()
    out["vacantes_simultaneas"] = out["empresa"].map(normalize).map(conteo).fillna(0).astype(int)
    return out


def fuzzy_dedup(out: pd.DataFrame, exact_sources) -> pd.DataFrame:
    """Colapsa vacantes con misma empresa+ciudad y título similar (thefuzz ≥85).
    Se conserva la versión de mayor prioridad (Computrabajo) y si las fuentes
    difieren la fuente queda como 'ambas'."""
    from thefuzz import fuzz

    # arrancar marcando "ambas" para las que ya colapsaron en el paso exacto
    def initial_fuente(row):
        srcs = exact_sources.get(row["_key"], {row["fuente"]}) - {""}
        real = {s for s in srcs if s in ("computrabajo", "indeed", "linkedin", "mock")}
        return "ambas" if len(real) > 1 else row["fuente"]

    out = out.copy()
    out["fuente"] = out.apply(initial_fuente, axis=1)
    out["_bucket"] = (out["empresa"].map(normalize).map(strip_accents) + "|" +
                      out["ciudad"].map(normalize).map(strip_accents))
    out["_prio"] = out["fuente"].map(lambda f: SOURCE_PRIORITY.get(f, 9))
    out = out.sort_values(["_bucket", "_prio"]).reset_index(drop=True)

    keep, kept_by_bucket = [], {}
    for idx, row in out.iterrows():
        bucket = row["_bucket"]
        title = strip_accents(normalize(row["titulo"]))
        duplicate_of = None
        for kept_idx in kept_by_bucket.get(bucket, []):
            kept_title = strip_accents(normalize(out.at[kept_idx, "titulo"]))
            score = max(fuzz.token_sort_ratio(title, kept_title),
                        fuzz.token_set_ratio(title, kept_title))
            if score >= 85:
                duplicate_of = kept_idx
                break
        if duplicate_of is None:
            keep.append(idx)
            kept_by_bucket.setdefault(bucket, []).append(idx)
        else:
            kept_fuente = out.at[duplicate_of, "fuente"]
            if row["fuente"] != kept_fuente or kept_fuente == "ambas":
                out.at[duplicate_of, "fuente"] = "ambas"
            # si la conservada no trae sueldo y la duplicada sí, complétalo
            for col in ("sueldo_min", "sueldo_max"):
                if pd.isna(out.at[duplicate_of, col]) and pd.notna(row[col]):
                    out.at[duplicate_of, col] = row[col]
    return out.loc[keep].drop(columns=["_bucket", "_prio"]).reset_index(drop=True)


# ----------------------------------------------------------------------------
# Excel
# ----------------------------------------------------------------------------

SHEET_COLS = {
    "detalle": [
        ("empresa", "Empresa", 32), ("titulo", "Título del puesto", 36),
        ("ciudad", "Ciudad", 20), ("estado", "Estado", 20),
        ("sueldo_min", "Sueldo min", 12), ("sueldo_max", "Sueldo max", 12),
        ("fecha_publicacion", "Fecha publicación", 16),
        ("tipo_contrato", "Tipo de contrato", 22),
        ("vacantes_simultaneas", "Vacantes simultáneas", 14),
        ("extracto", "Extracto de actividades", 60),
        ("palabras_senal", "Palabras de señal", 28),
        ("liga", "Liga", 40),
        ("keyword_busqueda", "Keyword de búsqueda", 24),
        ("plaza_busqueda", "Plaza de búsqueda", 22),
        ("fuente", "Fuente", 13),
        ("desc_completa", "Descripción completa obtenida", 14),
        ("agencia", "Agencia", 10),
        ("screenshot", "Captura", 30),
    ],
    "descartadas": [
        ("empresa", "Empresa", 32), ("titulo", "Título", 36),
        ("ciudad", "Ciudad", 20),
        ("motivo_descarte", "Motivo del descarte", 55),
        ("fuente", "Fuente", 13),
    ],
    "sin_verificar": [
        ("empresa", "Empresa", 32), ("titulo", "Título", 36),
        ("ciudad", "Ciudad", 20),
        ("liga", "Liga", 45),
        ("keyword_busqueda", "Keyword de búsqueda", 24),
        ("fuente", "Fuente", 13),
        ("desc_completa", "Descripción completa obtenida", 14),
    ],
}


def style_sheet(ws, n_cols: int, widths):
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="1F3864")  # azul oscuro
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    body_font = Font(name="Arial", size=10)

    for j in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=j)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(j)].width = widths[j - 1]
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


def write_df_sheet(writer, df: pd.DataFrame, sheet_name: str, spec_key: str):
    spec = SHEET_COLS[spec_key]
    cols = [c for c, _, _ in spec]
    headers = [h for _, h, _ in spec]
    widths = [w for _, _, w in spec]
    view = df.reindex(columns=cols)
    view.columns = headers
    view.to_excel(writer, sheet_name=sheet_name, index=False)
    style_sheet(writer.sheets[sheet_name], len(cols), widths)


def build_stats(processed: pd.DataFrame, raw_count: int) -> list:
    senal = processed[processed["clasificacion"] == "SEÑAL"]
    revisar = processed[processed["clasificacion"] == "REVISAR"]
    ruido = processed[processed["clasificacion"] == "RUIDO"]
    sin_ver = processed[processed["clasificacion"] == "SIN VERIFICAR"]

    rows = [
        ("Total vacantes crudas (antes de dedup)", raw_count),
        ("Total tras deduplicación (exacta + fuzzy)", len(processed)),
        ("Total SEÑAL", len(senal)),
        ("Total RUIDO (descartadas)", len(ruido)),
        ("Total REVISAR (casos borde)", len(revisar)),
        ("Total SIN VERIFICAR (reintentar)", len(sin_ver)),
        ("Empresas únicas con señal", senal["empresa"].map(normalize).nunique()),
    ]

    rows.append(("", ""))
    rows.append(("VACANTES POR FUENTE (tras dedup)", ""))
    for fuente, cnt in processed["fuente"].value_counts().items():
        rows.append((f"  {fuente}", int(cnt)))

    sueldos = pd.concat([senal, revisar])[["sueldo_min", "sueldo_max"]].apply(
        pd.to_numeric, errors="coerce")
    promedio = sueldos.stack().mean()
    rows.append(("Sueldo promedio publicado (señal + revisar)",
                 round(float(promedio), 2) if pd.notna(promedio) else "N/D"))

    rows.append(("", ""))
    rows.append(("TOP 10 EMPRESAS POR VACANTES SIMULTÁNEAS", ""))
    top = (senal.groupby("empresa")["vacantes_simultaneas"].max()
           .sort_values(ascending=False).head(10))
    for empresa, cnt in top.items():
        rows.append((empresa, int(cnt)))

    rows.append(("", ""))
    rows.append(("DISTRIBUCIÓN POR ESTADO (SEÑAL)", ""))
    for estado, cnt in senal["estado"].replace("", "(sin estado)").value_counts().items():
        rows.append((estado, int(cnt)))
    return rows


def write_excel(path: Path, processed: pd.DataFrame, raw_count: int):
    senal = processed[processed["clasificacion"] == "SEÑAL"]
    revisar = processed[processed["clasificacion"] == "REVISAR"]
    ruido = processed[processed["clasificacion"] == "RUIDO"]
    sin_ver = processed[processed["clasificacion"] == "SIN VERIFICAR"]

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        write_df_sheet(writer, senal, "Señal", "detalle")
        write_df_sheet(writer, revisar, "Revisar", "detalle")
        write_df_sheet(writer, ruido, "Descartadas", "descartadas")
        write_df_sheet(writer, sin_ver, "Sin verificar", "sin_verificar")

        stats = pd.DataFrame(build_stats(processed, raw_count),
                             columns=["Métrica", "Valor"])
        stats.to_excel(writer, sheet_name="Estadísticas", index=False)
        style_sheet(writer.sheets["Estadísticas"], 2, [55, 20])


def load_existing_keys(path: Path) -> set:
    """Llaves empresa|título|ciudad ya presentes en el Excel (para --append).
    'Sin verificar' se excluye a propósito: esas deben reintentar entrada."""
    if not path.exists():
        return set()
    keys = set()
    for sheet in ("Señal", "Revisar", "Descartadas"):
        try:
            df = pd.read_excel(path, sheet_name=sheet)
        except ValueError:
            continue
        emp = df.get("Empresa", pd.Series(dtype=str))
        tit = df.get("Título del puesto", df.get("Título", pd.Series(dtype=str)))
        ciu = df.get("Ciudad", pd.Series(dtype=str))
        for e, t, c in zip(emp, tit, ciu):
            keys.add(f"{strip_accents(normalize(e))}|{strip_accents(normalize(t))}|{strip_accents(normalize(c))}")
    return keys


def load_existing_rows(path: Path) -> pd.DataFrame:
    """Reconstruye el dataframe procesado desde un Excel existente (para --append)."""
    if not path.exists():
        return pd.DataFrame()
    frames = []
    for sheet, clasif, spec_key in (("Señal", "SEÑAL", "detalle"),
                                    ("Revisar", "REVISAR", "detalle"),
                                    ("Descartadas", "RUIDO", "descartadas")):
        try:
            df = pd.read_excel(path, sheet_name=sheet)
        except ValueError:
            continue
        inverse = {h: c for c, h, _ in SHEET_COLS[spec_key]}
        df = df.rename(columns=inverse)
        df["clasificacion"] = clasif
        frames.append(df)
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


# ----------------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Censo de vacantes de liquidación (Indeed MX + Computrabajo)")
    parser.add_argument("--hours-old", type=int, default=2160,
                        help="Antigüedad máxima en horas para Indeed (2160 = 90 días; 24 para diario)")
    parser.add_argument("--results-wanted", type=int, default=50)
    parser.add_argument("--append", action="store_true",
                        help="Agrega solo vacantes nuevas al Excel existente sin duplicar")
    parser.add_argument("--output", default="censo_liquidacion_indeed.xlsx")
    parser.add_argument("--raw-csv", default="respaldo_crudo.csv")
    parser.add_argument("--checkpoint-csv", default="respaldo_crudo_incremental.csv",
                        help="CSV incremental para reanudar corridas interrumpidas")
    parser.add_argument("--checkpoint-every", type=int, default=25)
    parser.add_argument("--capturas-dir", default="capturas")
    parser.add_argument("--storage-state", default="ct_storage_state.json")
    parser.add_argument("--min-sleep", type=float, default=3.0)
    parser.add_argument("--max-sleep", type=float, default=8.0)
    parser.add_argument("--headful", action="store_true",
                        help="Abre el navegador visible (para resolver captcha una vez)")
    parser.add_argument("--skip-computrabajo", action="store_true")
    parser.add_argument("--skip-indeed", action="store_true")
    parser.add_argument("--test", action="store_true",
                        help="Corre el pipeline con datos simulados (sin red)")
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
        handlers=[logging.StreamHandler(sys.stdout),
                  logging.FileHandler("censo_scraper.log", encoding="utf-8")],
    )

    started = datetime.now()
    output = Path(args.output)
    total_busquedas = len(KEYWORDS) * len(LOCATIONS) + len(KEYWORDS)  # jobspy + computrabajo

    if args.test:
        log.info("MODO PRUEBA: usando datos simulados, sin llamadas de red")
        raw, ok, failed = mock_jobs(), total_busquedas, 0
    else:
        checkpoint = Checkpoint(Path(args.checkpoint_csv), every=args.checkpoint_every)
        frames = []
        ok = failed = 0

        if not args.skip_indeed:
            jd, jok, jfail = run_jobspy_searches(
                args.results_wanted, args.hours_old,
                args.min_sleep, args.max_sleep, checkpoint)
            ok, failed = ok + jok, failed + jfail
            if len(jd) > 0:
                frames.append(jd)

        if not args.skip_computrabajo:
            ct = ComputrabajoScraper(
                headful=args.headful, storage_state=args.storage_state,
                capturas_dir=args.capturas_dir, seen_urls=checkpoint.seen_urls)
            try:
                ct.start()
                cd, cok, cfail = ct.run(KEYWORDS, checkpoint)
                ok, failed = ok + cok, failed + cfail
                if len(cd) > 0:
                    frames.append(cd)
            except Exception as exc:  # noqa: BLE001
                log.error("Computrabajo no pudo iniciar: %s", exc)
                failed += len(KEYWORDS)
            finally:
                ct.stop()

        checkpoint.flush()
        raw = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()

        # revisión una por una: completar descripciones truncadas de jobspy
        if len(raw) > 0:
            if "desc_verificada" not in raw.columns:
                raw["desc_verificada"] = False
            jobspy_mask = raw["fuente"].isin(["indeed", "linkedin"])
            raw.loc[jobspy_mask, "desc_verificada"] = (
                raw.loc[jobspy_mask, "description"].map(is_desc_complete))
            # las de computrabajo se visitaron individualmente: verificadas
            raw.loc[raw["fuente"] == "computrabajo", "desc_verificada"] = (
                raw.loc[raw["fuente"] == "computrabajo", "description"]
                .map(lambda d: len(normalize(d)) >= 100))
            if jobspy_mask.any():
                fixed = fetch_missing_descriptions(raw[jobspy_mask], args.headful)
                raw.loc[jobspy_mask, ["description", "desc_verificada"]] = (
                    fixed[["description", "desc_verificada"]].values)

    # respaldo crudo completo (para reprocesar sin volver a scrapear)
    if len(raw) > 0:
        if args.append and Path(args.raw_csv).exists():
            raw.to_csv(args.raw_csv, mode="a", header=False, index=False)
        else:
            raw.to_csv(args.raw_csv, index=False)
        log.info("Respaldo crudo guardado en %s (%d filas)", args.raw_csv, len(raw))

    processed = process(raw)

    nuevas = len(processed)
    if args.append and output.exists() and not processed.empty:
        existing_keys = load_existing_keys(output)
        processed = processed[~processed["_key"].isin(existing_keys)]
        nuevas = len(processed)
        log.info("Modo append: %d vacantes nuevas (no duplicadas)", nuevas)
        previas = load_existing_rows(output)
        if not previas.empty:
            processed = pd.concat([previas, processed], ignore_index=True)
            activas = processed[processed["clasificacion"].isin(["SEÑAL", "REVISAR"])]
            conteo = activas.groupby(activas["empresa"].map(normalize)).size()
            processed["vacantes_simultaneas"] = (
                processed["empresa"].map(normalize).map(conteo).fillna(0).astype(int))

    if processed.empty:
        log.warning("Sin vacantes que procesar; no se genera Excel")
    else:
        write_excel(output, processed, len(raw))
        log.info("Excel generado: %s", output)

    def count(clasif):
        return int((processed["clasificacion"] == clasif).sum()) if not processed.empty else 0

    print("\n" + "=" * 62)
    print("RESUMEN DEL CENSO")
    print("=" * 62)
    print(f"Búsquedas corridas:         {ok + failed} de {total_busquedas}")
    print(f"Búsquedas fallidas:         {failed}")
    print(f"Vacantes crudas:            {len(raw)}")
    print(f"Tras deduplicación:         {len(processed)}")
    print(f"Pasaron el filtro (SEÑAL):  {count('SEÑAL')}")
    print(f"Casos borde (REVISAR):      {count('REVISAR')}")
    print(f"Descartadas (RUIDO):        {count('RUIDO')}")
    print(f"Sin verificar (reintentar): {count('SIN VERIFICAR')}")
    if args.append:
        print(f"Vacantes nuevas agregadas:  {nuevas}")
    print(f"Duración: {datetime.now() - started}")
    print("=" * 62)


if __name__ == "__main__":
    main()
