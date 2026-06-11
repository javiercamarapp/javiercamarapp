#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generar_quiniela_final.py — Fase 3A
====================================
Copia el template reconstruido, escribe SOLO marcadores (cols D/E de Grupos y
G/H de Bracket), resuelve los slots de terceros con la simulación, llena
Bota de Oro y último lugar, y VERIFICA con recálculo LibreOffice que:
  - posiciones auto-calculadas == simulación Python
  - terceros top-8 == simulación
  - ganadores W## del bracket == simulación
  - KO sin empates y campeón coherente
Output: entregables/QUINIELA_2026_JAVIER_FINAL.xlsx
"""

import json
import os
import shutil
import subprocess
import sys

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PRED = json.load(open(os.path.join(BASE, "data", "predicciones_104.json"), encoding="utf-8"))

from openpyxl import load_workbook

BOTA_DE_ORO = "Kylian Mbappé"      # favorito +600 (data/momios_mx.json); Francia llega a la final
ULTIMO_LUGAR = "Curazao"            # 0 pts, -5 dif en simulación; Elo 1500 (3° más bajo); ref. previa

TEMPLATE = os.path.join(BASE, "QUINIELA_2026_TEMPLATE_RECONSTRUIDO.xlsx")
SALIDA = os.path.join(BASE, "entregables", "QUINIELA_2026_JAVIER_FINAL.xlsx")


def llenar():
    shutil.copy(TEMPLATE, SALIDA)
    wb = load_workbook(SALIDA)
    ws = wb["Grupos"]
    # mapear num de partido -> fila escaneando la columna A
    fila_de = {}
    for row in ws.iter_rows(min_col=1, max_col=1):
        c = row[0]
        if isinstance(c.value, int):
            fila_de[c.value] = c.row
    grupos = [p for p in PRED["predicciones"] if p["fase"] == "grupos"]
    assert len(grupos) == 72
    for p in grupos:
        r = fila_de[p["num"]]
        assert ws.cell(r, 3).value == p["local"], (p["num"], ws.cell(r, 3).value, p["local"])
        ws.cell(r, 4, p["gl"])
        ws.cell(r, 5, p["gv"])

    wsb = wb["Bracket"]
    fila_b = {}
    for row in wsb.iter_rows(min_col=1, max_col=1):
        c = row[0]
        if isinstance(c.value, int):
            fila_b[c.value] = c.row
    ko = [p for p in PRED["predicciones"] if p["fase"] != "grupos"]
    assert len(ko) == 32
    for p in ko:
        r = fila_b[p["num"]]
        gl, gv = p["gl"], p["gv"]
        if gl == gv:
            raise SystemExit(f"KO con empate en #{p['num']}")
        wsb.cell(r, 7, gl)
        wsb.cell(r, 8, gv)
        # slots de terceros: escribir el equipo como VALOR
        for col_slot, col_eq, slot_key in ((5, 6, "slot_local"), (10, 9, "slot_visitante")):
            slot = p[slot_key]
            if slot.startswith("3_"):
                equipo = p["local"] if slot_key == "slot_local" else p["visitante"]
                wsb.cell(r, col_eq, equipo)

    wse = wb["Extras"]
    wse["B3"] = BOTA_DE_ORO
    wse["B4"] = ULTIMO_LUGAR
    wb.save(SALIDA)
    print(f"Quiniela escrita: {SALIDA}")


def verificar():
    outdir = "/tmp/recalc_final"
    subprocess.run(["soffice", "--headless", "-env:UserInstallation=file:///tmp/loprofile",
                    "--convert-to", "xlsx", "--outdir", outdir, SALIDA],
                   check=True, capture_output=True, timeout=300)
    wb = load_workbook(os.path.join(outdir, os.path.basename(SALIDA)), data_only=True)
    ws = wb["Grupos"]
    fallos = []

    # 1. posiciones por grupo vs simulación
    bloque = {}
    for row in ws.iter_rows(min_col=1, max_col=1):
        v = row[0].value
        if isinstance(v, str) and v.startswith("GRUPO "):
            bloque[v.split()[1]] = row[0].row + 2
    for g, tabla in PRED["tablas_grupos"].items():
        r0 = bloque[g]
        excel = sorted(((ws.cell(r0 + i, 8).value, ws.cell(r0 + i, 15).value) for i in range(4)),
                       key=lambda t: t[1])
        esperado = [f["equipo"] for f in tabla]
        if [e for e, _ in excel] != esperado:
            fallos.append(f"Grupo {g}: Excel {[e for e,_ in excel]} != sim {esperado}")

    # 2. terceros top-8
    wst = wb["Terceros"]
    excel_terceros = []
    for i in range(12):
        r = 4 + i
        excel_terceros.append((wst.cell(r, 2).value, wst.cell(r, 7).value, wst.cell(r, 8).value))
    avanzan_excel = {e for e, rank, av in excel_terceros if av == "SÍ"}
    avanzan_sim = {t["equipo"] for t in PRED["terceros_ranking"][:8]}
    if avanzan_excel != avanzan_sim:
        fallos.append(f"Terceros: Excel {sorted(avanzan_excel)} != sim {sorted(avanzan_sim)}")

    # 3. ganadores del bracket y campeón
    wsb = wb["Bracket"]
    ganador_sim = {p["num"]: p["ganador"] for p in PRED["predicciones"] if p["fase"] != "grupos"}
    campeon = None
    for row in wsb.iter_rows(min_col=1, max_col=10):
        num = row[0].value
        if isinstance(num, int) and num >= 73:
            gl, gv = row[6].value, row[7].value
            eq_l, eq_v = row[5].value, row[8].value
            if gl == gv:
                fallos.append(f"#{num}: empate {gl}-{gv}")
                continue
            w = eq_l if gl > gv else eq_v
            if w != ganador_sim[num]:
                fallos.append(f"#{num}: Excel ganador {w} != sim {ganador_sim[num]} ({eq_l} {gl}-{gv} {eq_v})")
            if num == 104:
                campeon = w
    print(f"Campeón según Excel recalculado: {campeon}")
    if campeon != ganador_sim[104]:
        fallos.append(f"Campeón Excel {campeon} != sim {ganador_sim[104]}")

    if fallos:
        print("FALLOS:")
        for f in fallos:
            print("  ✗", f)
        sys.exit(1)
    print("PASS: 72 marcadores + posiciones + terceros + bracket + campeón coherentes (recalc LibreOffice)")


if __name__ == "__main__":
    os.makedirs(os.path.join(BASE, "entregables"), exist_ok=True)
    llenar()
    verificar()
