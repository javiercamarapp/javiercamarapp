#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generar_template.py — Genera QUINIELA_2026_TEMPLATE_RECONSTRUIDO.xlsx
======================================================================
Template funcional equivalente al de la quiniela de Pablo (el original no
existía en el repo). Hoja "Grupos": 72 marcadores editables + posiciones
AUTO-CALCULADAS por fórmulas (pts → dif → GF → alfabético). Hoja "Terceros":
ranking de los 12 terceros por fórmulas. Hoja "Bracket": partidos 73-104;
los slots 1X/2X y los ganadores W## se resuelven por FÓRMULAS; los slots de
terceros (3_XXXXX) los escribe el script de simulación como VALOR (la tabla
FIFA de 495 combinaciones no es practicable en fórmulas), etiquetado.

Uso: python3 scripts/generar_template.py [salida.xlsx]
"""

import json
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CAL = json.load(open(os.path.join(BASE, "data", "calendario.json"), encoding="utf-8"))
BRACKET = json.load(open(os.path.join(BASE, "data", "bracket.json"), encoding="utf-8"))

GRUPOS = "ABCDEFGHIJKL"

AZUL = PatternFill("solid", fgColor="1F4E79")
GRIS = PatternFill("solid", fgColor="D9E1F2")
AMARILLO = PatternFill("solid", fgColor="FFF2CC")
NEGRITA_BLANCA = Font(bold=True, color="FFFFFF")
NEGRITA = Font(bold=True)
CENTRO = Alignment(horizontal="center")


def equipos_de_grupo(g):
    """Equipos del grupo en orden de aparición en el calendario."""
    eqs = []
    for p in sorted((p for p in CAL if p["grupo"] == g), key=lambda x: x["num"]):
        for e in (p["local"], p["visitante"]):
            if e not in eqs:
                eqs.append(e)
    assert len(eqs) == 4, f"grupo {g}: {eqs}"
    return eqs


def hoja_grupos(wb):
    ws = wb.active
    ws.title = "Grupos"
    ws["A1"] = "QUINIELA MUNDIAL 2026 — FASE DE GRUPOS (llena SOLO las columnas D y E)"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = ("Posiciones AUTO-CALCULADAS: pts → dif → GF → alfabético. NO editar columnas H-Q. "
                "Template reconstruido 11-jun-2026 (ver data/estructura_torneo.md).")
    fila = 4
    bloques = {}  # grupo -> dict con rangos de celdas
    for g in GRUPOS:
        partidos = sorted((p for p in CAL if p["grupo"] == g), key=lambda x: x["num"])
        eqs = equipos_de_grupo(g)
        ws.cell(fila, 1, f"GRUPO {g}").font = NEGRITA
        ws.cell(fila, 1).fill = GRIS
        for c, t in enumerate(["#", "Fecha", "Local", "GL", "GV", "Visitante"], start=1):
            cel = ws.cell(fila + 1, c, t)
            cel.font = NEGRITA_BLANCA
            cel.fill = AZUL
            cel.alignment = CENTRO
        for c, t in enumerate(["Equipo", "PJ", "Pts", "GF", "GC", "Dif", "Score", "Pos"], start=8):
            cel = ws.cell(fila + 1, c, t)
            cel.font = NEGRITA_BLANCA
            cel.fill = AZUL
            cel.alignment = CENTRO
        r0 = fila + 2          # primera fila de partidos
        r5 = r0 + 5            # última
        for i, p in enumerate(partidos):
            r = r0 + i
            ws.cell(r, 1, p["num"])
            ws.cell(r, 2, p["fecha"])
            ws.cell(r, 3, p["local"])
            ws.cell(r, 4).fill = AMARILLO
            ws.cell(r, 5).fill = AMARILLO
            ws.cell(r, 6, p["visitante"])
        rangoC, rangoD = f"$C${r0}:$C${r5}", f"$D${r0}:$D${r5}"
        rangoE, rangoF = f"$E${r0}:$E${r5}", f"$F${r0}:$F${r5}"
        for i, eq in enumerate(eqs):
            r = r0 + i
            t = f"$H{r}"
            ws.cell(r, 8, eq)
            lleno = f"({rangoD}<>\"\")*({rangoE}<>\"\")"
            ws.cell(r, 9, (f"=SUMPRODUCT(({rangoC}={t})*{lleno})"
                           f"+SUMPRODUCT(({rangoF}={t})*{lleno})"))
            ws.cell(r, 10, (f"=SUMPRODUCT(({rangoC}={t})*{lleno}*({rangoD}>{rangoE}))*3"
                            f"+SUMPRODUCT(({rangoF}={t})*{lleno}*({rangoE}>{rangoD}))*3"
                            f"+SUMPRODUCT((({rangoC}={t})+({rangoF}={t}))*{lleno}*({rangoD}={rangoE}))"))
            ws.cell(r, 11, (f"=SUMPRODUCT(({rangoC}={t})*{lleno}*{rangoD})"
                            f"+SUMPRODUCT(({rangoF}={t})*{lleno}*{rangoE})"))
            ws.cell(r, 12, (f"=SUMPRODUCT(({rangoC}={t})*{lleno}*{rangoE})"
                            f"+SUMPRODUCT(({rangoF}={t})*{lleno}*{rangoD})"))
            ws.cell(r, 13, f"=K{r}-L{r}")
            # Score: pts*1e7 + (dif+100)*1e4 + GF*10 + desempate alfabético (3-#nombres_menores)
            ws.cell(r, 14, (f"=J{r}*10000000+(M{r}+100)*10000+K{r}*10"
                            f"+(3-COUNTIF($H${r0}:$H${r5},\"<\"&H{r}))"))
            ws.cell(r, 15, f"=RANK(N{r},$N${r0}:$N${r5})")
        # celdas 1°,2°,3° del grupo
        ws.cell(r0, 17, "1°:")
        ws.cell(r0, 18, f"=INDEX($H${r0}:$H${r5},MATCH(1,$O${r0}:$O${r5},0))")
        ws.cell(r0 + 1, 17, "2°:")
        ws.cell(r0 + 1, 18, f"=INDEX($H${r0}:$H${r5},MATCH(2,$O${r0}:$O${r5},0))")
        ws.cell(r0 + 2, 17, "3°:")
        ws.cell(r0 + 2, 18, f"=INDEX($H${r0}:$H${r5},MATCH(3,$O${r0}:$O${r5},0))")
        bloques[g] = {"r0": r0, "r5": r5,
                      "pos1": f"Grupos!$R${r0}", "pos2": f"Grupos!$R${r0 + 1}", "pos3": f"Grupos!$R${r0 + 2}",
                      "fila_tercero_stats": r0}
        fila = r5 + 2
    anchos = {1: 5, 2: 11, 3: 20, 4: 5, 5: 5, 6: 20, 8: 20, 9: 5, 10: 5, 11: 5,
              12: 5, 13: 6, 14: 13, 15: 5, 17: 4, 18: 20}
    for c, w in anchos.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    return bloques


def hoja_terceros(wb, bloques):
    ws = wb.create_sheet("Terceros")
    ws["A1"] = "RANKING DE TERCEROS LUGARES (auto-calculado; los 8 primeros avanzan)"
    ws["A1"].font = Font(bold=True, size=12)
    for c, t in enumerate(["Grupo", "Equipo", "Pts", "Dif", "GF", "Score", "Rank", "¿Avanza?"], start=1):
        cel = ws.cell(3, c, t)
        cel.font = NEGRITA_BLANCA
        cel.fill = AZUL
    r0, r5 = 4, 4 + 11
    for i, g in enumerate(GRUPOS):
        r = r0 + i
        b = bloques[g]
        gr0, gr5 = b["r0"], b["r5"]
        ws.cell(r, 1, g)
        ws.cell(r, 2, f"=INDEX(Grupos!$H${gr0}:$H${gr5},MATCH(3,Grupos!$O${gr0}:$O${gr5},0))")
        ws.cell(r, 3, f"=INDEX(Grupos!$J${gr0}:$J${gr5},MATCH(3,Grupos!$O${gr0}:$O${gr5},0))")
        ws.cell(r, 4, f"=INDEX(Grupos!$M${gr0}:$M${gr5},MATCH(3,Grupos!$O${gr0}:$O${gr5},0))")
        ws.cell(r, 5, f"=INDEX(Grupos!$K${gr0}:$K${gr5},MATCH(3,Grupos!$O${gr0}:$O${gr5},0))")
        ws.cell(r, 6, (f"=C{r}*10000000+(D{r}+100)*10000+E{r}*10"
                       f"+(11-COUNTIF($B${r0}:$B${r5},\"<\"&B{r}))"))
        ws.cell(r, 7, f"=RANK(F{r},$F${r0}:$F${r5})")
        ws.cell(r, 8, f"=IF(G{r}<=8,\"SÍ\",\"no\")")
    for c, w in {1: 7, 2: 20, 3: 5, 4: 6, 5: 5, 6: 13, 7: 6, 8: 9}.items():
        ws.column_dimensions[get_column_letter(c)].width = w


def hoja_bracket(wb, bloques):
    ws = wb.create_sheet("Bracket")
    ws["A1"] = "ELIMINATORIA (partidos 73-104) — llena SOLO GL/GV; SIN empates (penales = +1 gol al ganador)"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = ("Slots 1X/2X y ganadores W## por fórmula. Slots de terceros (3°): los escribe la simulación "
                "como VALOR (asignación FIFA aproximada por elegibilidad — ver data/bracket.json).")
    for c, t in enumerate(["#", "Ronda", "Fecha", "Ciudad", "Slot L", "Local", "GL", "GV", "Visitante", "Slot V"], start=1):
        cel = ws.cell(4, c, t)
        cel.font = NEGRITA_BLANCA
        cel.fill = AZUL
    rondas = ([("Dieciseisavos", p) for p in BRACKET["dieciseisavos"]]
              + [("Octavos", p) for p in BRACKET["octavos"]]
              + [("Cuartos", p) for p in BRACKET["cuartos"]]
              + [("Semifinal", p) for p in BRACKET["semifinales"]]
              + [("3er lugar", BRACKET["tercer_lugar"]), ("FINAL", BRACKET["final"])])
    fila_de = {}  # num partido -> fila
    r = 5
    for ronda, p in rondas:
        fila_de[p["num"]] = r
        ws.cell(r, 1, p["num"])
        ws.cell(r, 2, ronda)
        ws.cell(r, 3, p["fecha"])
        ws.cell(r, 4, p["ciudad"])
        ws.cell(r, 7).fill = AMARILLO
        ws.cell(r, 8).fill = AMARILLO
        for col_slot, col_eq, slot in ((5, 6, p["local"]), (10, 9, p["visitante"])):
            ws.cell(r, col_slot, slot)
            if slot.startswith(("1", "2")) and len(slot) == 2:
                g = slot[1]
                b = bloques[g]
                fila_pos = b["r0"] + int(slot[0]) - 1
                ws.cell(r, col_eq, f"=Grupos!$R${fila_pos}")
            elif slot.startswith("3_"):
                ws.cell(r, col_eq, "(simulación)")  # la llena el script final como VALOR
            elif slot.startswith(("W", "L")):
                rf = fila_de[int(slot[1:])]
                if slot.startswith("W"):
                    ws.cell(r, col_eq, f"=IF(G{rf}>H{rf},F{rf},I{rf})")
                else:
                    ws.cell(r, col_eq, f"=IF(G{rf}>H{rf},I{rf},F{rf})")
        r += 1
    ws.cell(r + 1, 1, "CAMPEÓN:").font = NEGRITA
    rf = fila_de[104]
    ws.cell(r + 1, 2, f"=IF(G{rf}>H{rf},F{rf},I{rf})")
    for c, w in {1: 5, 2: 13, 3: 11, 4: 16, 5: 9, 6: 20, 7: 5, 8: 5, 9: 20, 10: 9}.items():
        ws.column_dimensions[get_column_letter(c)].width = w


def hoja_extras(wb):
    ws = wb.create_sheet("Extras")
    ws["A1"] = "PREGUNTAS EXTRA (25 pts c/u)"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A3"] = "Bota de Oro (máximo goleador):"
    ws["B3"].fill = AMARILLO
    ws["A4"] = "Último lugar del torneo:"
    ws["B4"].fill = AMARILLO
    ws["A6"] = "DATOS DEL PARTICIPANTE (dejar en blanco)"
    ws["A6"].font = NEGRITA
    for i, campo in enumerate(["Nombre:", "Email:", "Teléfono:"], start=7):
        ws.cell(i, 1, campo)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 28


def main():
    salida = sys.argv[1] if len(sys.argv) > 1 else os.path.join(BASE, "QUINIELA_2026_TEMPLATE_RECONSTRUIDO.xlsx")
    wb = Workbook()
    bloques = hoja_grupos(wb)
    # columna R de la hoja Grupos guarda 1°/2°/3° (ya escrita en hoja_grupos col 18=R)
    hoja_terceros(wb, bloques)
    hoja_bracket(wb, bloques)
    hoja_extras(wb)
    wb.save(salida)
    print(f"Template generado: {salida}")


if __name__ == "__main__":
    main()
